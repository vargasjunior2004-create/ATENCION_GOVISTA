from io import BytesIO
from pathlib import Path

from django.core.signing import TimestampSigner, BadSignature, SignatureExpired

from .models import Sale, CashCount, Outflow

SIGNER = TimestampSigner()

REQUEST_TYPE_LABELS = {
    'nuevo_contrato': 'INSTALACIONES',
    'cambio_plan': 'CAMBIO DE PLAN',
    'recontratacion': 'RECONTRATACION',
    'retiro': 'RETIROS',
    'adicion': 'ADICION',
    'baja_temporal': 'BAJA TEMPORAL',
    'otro': 'OTROS',
}

LOGO_PATH = Path(__file__).resolve().parent / 'logo.png'

DENOMINATIONS = [
    ('Moneda 0.50', 0.50), ('Moneda 1', 1), ('Moneda 2', 2), ('Moneda 5', 5),
    ('Billete 10', 10), ('Billete 20', 20), ('Billete 50', 50),
    ('Billete 100', 100), ('Billete 200', 200),
]


def sign_report_token(payload):
    """Firma 'from:to' (o 'date') con expiración de 1 hora."""
    return SIGNER.sign(payload)


def unsign_report_token(token):
    """Valida la firma. Devuelve el payload o None si es inválida/expirada."""
    try:
        return SIGNER.unsign(token, max_age=3600)
    except (BadSignature, SignatureExpired):
        return None


def _logo_image(max_width=50):
    """Devuelve un flowable Image con el logo de la empresa, o None si no existe."""
    try:
        from reportlab.lib.utils import ImageReader
        from reportlab.platypus import Image

        if not LOGO_PATH.exists():
            return None
        width, height = ImageReader(str(LOGO_PATH)).getSize()
        scale = max_width / width
        return Image(str(LOGO_PATH),
                     width=max_width, height=height * scale,
                     hAlign='CENTER')
    except Exception:
        return None


def _report_header(title, subtitle):
    """Logo + título + subtítulo para encabezar un reporte."""
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Spacer

    styles = getSampleStyleSheet()
    story = []
    logo = _logo_image()
    if logo:
        story.append(logo)
        story.append(Spacer(1, 6))
    story.append(title)
    story.append(Spacer(1, 4))
    story.append(subtitle)
    story.append(Spacer(1, 10))
    return story


# ---------------------------------------------------------------- PDF (sales)

def build_sales_pdf(from_date, to_date, request_type=None, service_type=None):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet

    sales = Sale.objects.select_related('plan', 'createdBy').filter(
        date__gte=from_date, date__lte=to_date).order_by('date', 'id')
    if request_type:
        sales = sales.filter(requestType=request_type)
    if service_type:
        sales = sales.filter(serviceType=service_type)

    # Service type mapping
    service_type_map = {
        'internet': 'INTERNET',
        'tv': 'TV ANALOGA',
        'tv_digital': 'TV DIGITAL',
        'combo_analog': 'INTERNET + TV ANALOGA',
        'combo_digital': 'INTERNET + TV DIGITAL',
    }

    styles = getSampleStyleSheet()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)

    SERVICE_TYPE_LABELS = {
        'internet': 'INTERNET', 'tv': 'TV ANALOGA', 'tv_digital': 'TV DIGITAL',
        'combo_analog': 'INTERNET + TV ANALOGA', 'combo_digital': 'INTERNET + TV DIGITAL',
    }
    subtitle_parts = []
    if request_type:
        subtitle_parts.append(REQUEST_TYPE_LABELS.get(request_type, request_type.upper()))
    if service_type:
        subtitle_parts.append(SERVICE_TYPE_LABELS.get(service_type, service_type.upper()))
    title_label = ' - '.join(subtitle_parts) if subtitle_parts else 'TODOS'

    story = _report_header(
        Paragraph(f'MOV. CLIENTES - {title_label}', styles['Title']),
        Paragraph(f'Periodo: {from_date} al {to_date}', styles['Normal']),
    )

    if request_type == 'retiro':
        header = ['Fecha', 'Kardex', 'Cliente', 'Servicio', 'Solicitud', 'Plan', 'Motivo', 'Monto Ini', 'Dif', 'Operador']
    else:
        header = ['Fecha', 'Kardex', 'Cliente', 'Servicio', 'Solicitud', 'Plan', 'Monto Ini', 'Dif', 'Operador']
    rows = [header]
    for s in sales:
        service_label = service_type_map.get(s.serviceType, s.serviceType)
        if request_type == 'retiro':
            rows.append([
                s.date.strftime('%d/%m/%Y') if s.date else '',
                s.clientCode, s.clientName,
                service_label, s.get_requestType_display(),
                s.plan.label,
                s.changeReason or '-',
                f'{float(s.total):.2f}',
                f'{float(s.total):.2f}',
                s.createdBy.name,
            ])
        else:
            rows.append([
                s.date.strftime('%d/%m/%Y') if s.date else '',
                s.clientCode, s.clientName,
                service_label, s.get_requestType_display(),
                s.plan.label,
                f'{float(s.total):.2f}',
                f'{float(s.total):.2f}',
                s.createdBy.name,
            ])

    table = Table(rows, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1d4ed8')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    story.append(table)
    story.append(Spacer(1, 6))
    story.append(Paragraph(f'{len(sales)} registros en el periodo', styles['Normal']))
    doc.build(story)
    buf.seek(0)
    return buf


# --------------------------------------------------------------- PDF (caja)

def build_cash_pdf(date_str):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    cc = CashCount.objects.filter(date=date_str).first()
    outflows = list(Outflow.objects.filter(date=date_str))
    total_out = sum(float(o.amount) for o in outflows)

    styles = getSampleStyleSheet()
    
    # Custom styles
    styles['Title'].fontSize = 16
    styles['Title'].alignment = TA_CENTER
    
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=15 * mm, rightMargin=15 * mm,
                            topMargin=15 * mm, bottomMargin=15 * mm)

    registered_by = cc.createdBy.name if cc and cc.createdBy else None
    
    # Generate arqueo number based on date
    from datetime import datetime
    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
    arqueo_num = f"{date_obj.strftime('%Y%m%d')}-001"
    
    story = []
    
    # Header section
    header_data = [
        ['FECHA:', date_obj.strftime('%d/%m/%Y'), 'CAJA:', 'GO VISTA', 'ARQUEO N°:', arqueo_num],
        ['CAJERO:', registered_by or '—', '', '', '', ''],
    ]
    header_table = Table(header_data, colWidths=[20*mm, 35*mm, 15*mm, 40*mm, 25*mm, 35*mm])
    header_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 10))
    
    # Saldo section
    saldo_data = [
        ['SALDO', ''],
        ['Saldo Inicial', '0.00'],
    ]
    saldo_table = Table(saldo_data, colWidths=[60*mm, 40*mm])
    saldo_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e5e7eb')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
    ]))
    story.append(saldo_table)
    story.append(Spacer(1, 10))
    
    # Efectivo section
    story.append(Paragraph('EFECTIVO - EQUIVALENTE DE EFECTIVO', styles['Heading2']))
    story.append(Spacer(1, 6))
    
    if cc:
        # Calculate totals
        coins_total = (cc.coin_050 * 0.50 + cc.coin_1 * 1 + cc.coin_2 * 2 + cc.coin_5 * 5)
        bills_total = (cc.bill_10 * 10 + cc.bill_20 * 20 + cc.bill_50 * 50 + 
                      cc.bill_100 * 100 + cc.bill_200 * 200)
        
        # Coins section
        coins_data = [
            ['MONEDAS', '', f'{coins_total:.2f}'],
            ['Denominación', 'Cantidad', 'Total'],
            ['0.50', str(cc.coin_050), f'{cc.coin_050 * 0.50:.2f}'],
            ['1.00', str(cc.coin_1), f'{cc.coin_1 * 1:.2f}'],
            ['2.00', str(cc.coin_2), f'{cc.coin_2 * 2:.2f}'],
            ['5.00', str(cc.coin_5), f'{cc.coin_5 * 5:.2f}'],
        ]
        coins_table = Table(coins_data, colWidths=[40*mm, 30*mm, 40*mm])
        coins_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e5e7eb')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
        ]))
        story.append(coins_table)
        story.append(Spacer(1, 10))
        
        # Bills section
        bills_data = [
            ['BILLETES', '', f'{bills_total:.2f}'],
            ['Denominación', 'Cantidad', 'Total'],
            ['10.00', str(cc.bill_10), f'{cc.bill_10 * 10:.2f}'],
            ['20.00', str(cc.bill_20), f'{cc.bill_20 * 20:.2f}'],
            ['50.00', str(cc.bill_50), f'{cc.bill_50 * 50:.2f}'],
            ['100.00', str(cc.bill_100), f'{cc.bill_100 * 100:.2f}'],
            ['200.00', str(cc.bill_200), f'{cc.bill_200 * 200:.2f}'],
        ]
        bills_table = Table(bills_data, colWidths=[40*mm, 30*mm, 40*mm])
        bills_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e5e7eb')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
        ]))
        story.append(bills_table)
        story.append(Spacer(1, 10))
        
        # Checks section (empty for now)
        checks_data = [
            ['CHEQUES', '', '-'],
            ['BCP', '', ''],
        ]
        checks_table = Table(checks_data, colWidths=[40*mm, 30*mm, 40*mm])
        checks_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e5e7eb')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
        ]))
        story.append(checks_table)
        story.append(Spacer(1, 10))
        
        # Other items section (outflows grouped by person)
        otros_data = [
            ['OTROS (FACTURAS/DEVOLUCIONES)', '', f'{total_out:.2f}'],
        ]
        # Group outflows by person
        person_totals = {}
        for o in outflows:
            person = o.personName
            if person not in person_totals:
                person_totals[person] = 0
            person_totals[person] += float(o.amount)
        
        for person, amount in person_totals.items():
            otros_data.append([person, '', f'{amount:.2f}'])
        
        otros_table = Table(otros_data, colWidths=[40*mm, 30*mm, 40*mm])
        otros_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f3f4f6')),
            ('FONTNAME', (0, 1), (-1, 1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#e5e7eb')),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (1, 0), (2, -1), 'RIGHT'),
        ]))
        story.append(otros_table)
        story.append(Spacer(1, 15))
        
        # Totals section
        total_caja = float(cc.total) + total_out
        totals_data = [
            ['', '', '1.- TOTAL CAJA CHICA', f'{total_caja:.2f}'],
            ['', '', '3.- TOTAL EFECTIVO', f'{float(cc.total):.2f}'],
        ]
        totals_table = Table(totals_data, colWidths=[30*mm, 30*mm, 50*mm, 40*mm])
        totals_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('ALIGN', (2, 0), (3, -1), 'RIGHT'),
            ('LINEBELOW', (2, 0), (3, 0), 1, colors.black),
            ('LINEBELOW', (2, 1), (3, 1), 1, colors.black),
        ]))
        story.append(totals_table)
        story.append(Spacer(1, 20))
        
        # Audit section
        audit_data = [
            ['AUDITOR', '________________________', '', 'SALDO _________________'],
        ]
        audit_table = Table(audit_data, colWidths=[30*mm, 60*mm, 20*mm, 50*mm])
        audit_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
        ]))
        story.append(audit_table)
        story.append(Spacer(1, 15))
        
        # Final total
        story.append(Paragraph(f'TOTAL: {total_caja:.2f} Bs', styles['Title']))
    else:
        story.append(Paragraph('No hay arqueo registrado para esta fecha.',
                               styles['Normal']))
    
    doc.build(story)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------- XLSX (sales)

def build_sales_xlsx(from_date, to_date):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    sales = Sale.objects.select_related('plan', 'createdBy').filter(
        date__gte=from_date, date__lte=to_date).order_by('date', 'id')

    wb = Workbook()
    ws = wb.active
    ws.title = 'MOV. CLIENTES'

    # Encabezados exactos como en el Excel original
    headers = [
        'FECHA', 'KARDEX', 'NOMBRE CLIENTE', 'TIPO DE SERVICIO', 'TIPO DE SOLICITUD',
        'PAQUETE TV CABLE', 'PAQUETE INTERNET', 'MONTO INICIAL', 'DIFERENCIA',
        'CAJERA(O)', 'MOTIVO CAMBIO DE PLAN',
        'PAQUETE CAMBIO TV CABLE', 'PAQUETE CAMBIO INTERNET', 'COMENTARIOS'
    ]

    # Estilos
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='1D4ED8')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Escribir encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Service type mapping
    service_type_map = {
        'internet': 'INTERNET',
        'tv': 'TV ANALOGA',
        'tv_digital': 'TV DIGITAL',
        'combo_analog': 'INTERNET + TV ANALOGA',
        'combo_digital': 'INTERNET + TV DIGITAL',
    }

    # Escribir datos
    for row_idx, s in enumerate(sales, 2):
        service_label = service_type_map.get(s.serviceType, s.serviceType)
        
        # Determine package TV and Internet based on service type
        paq_tv = ''
        paq_inet = ''
        if 'tv' in s.serviceType or 'combo' in s.serviceType:
            paq_tv = s.plan.label
        if 'internet' in s.serviceType or 'combo' in s.serviceType:
            paq_inet = s.plan.label

        data_row = [
            s.date.strftime('%d/%m/%Y') if s.date else '',
            s.clientCode,
            s.clientName,
            service_label,
            s.get_requestType_display(),
            paq_tv,
            paq_inet,
            float(s.total),
            float(s.total),
            s.createdBy.name,
            s.changeReason if s.requestType == 'cambio_plan' else '',
            '',  # Paquete cambio TV (no aplica por ahora)
            '',  # Paquete cambio Internet (no aplica por ahora)
            s.notes,
        ]

        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col in [8, 9]:  # Montos
                cell.number_format = '#,##0.00'

    # Anchos de columna
    column_widths = [12, 12, 28, 24, 18, 18, 18, 14, 14, 20, 20, 20, 20, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else 'A' + chr(64 + i - 26)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------- PNG (sales)

def build_sales_png(from_date, to_date):
    from PIL import Image, ImageDraw, ImageFont

    sales = Sale.objects.select_related('plan', 'createdBy').filter(
        date__gte=from_date, date__lte=to_date).order_by('date', 'id')

    service_type_map = {
        'internet': 'INTERNET',
        'tv': 'TV ANALOGA',
        'tv_digital': 'TV DIGITAL',
        'combo_analog': 'INTERNET + TV ANALOGA',
        'combo_digital': 'INTERNET + TV DIGITAL',
    }

    headers = ['FECHA', 'KARDEX', 'CLIENTE', 'SERVICIO', 'SOLICITUD', 'PLAN', 'MONTO', 'CAJERA']
    col_widths = [140, 130, 300, 240, 180, 180, 120, 240]
    row_height = 40
    padding = 16

    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 14)
        font_bold = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 14)
        font_title = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 20)
    except Exception:
        font = ImageFont.load_default()
        font_bold = font
        font_title = font

    total_width = sum(col_widths) + padding * 2
    title_height = 60
    total_height = title_height + row_height + row_height * len(sales) + 40

    img = Image.new('RGB', (total_width, total_height), '#FFFFFF')
    draw = ImageDraw.Draw(img)

    # Title
    draw.text((padding, 15), f'MOV. CLIENTES  {from_date} al {to_date}', fill='#1D4ED8', font=font_title)

    # Header row
    y = title_height
    draw.rectangle([0, y, total_width, y + row_height], fill='#1D4ED8')
    x = padding
    for i, h in enumerate(headers):
        draw.text((x + 6, y + 10), h, fill='#FFFFFF', font=font_bold)
        x += col_widths[i]

    # Data rows
    y += row_height
    for idx, s in enumerate(sales):
        bg = '#F8FAFC' if idx % 2 == 0 else '#FFFFFF'
        draw.rectangle([0, y, total_width, y + row_height], fill=bg)
        service_label = service_type_map.get(s.serviceType, s.serviceType)
        values = [
            s.date.strftime('%d/%m/%Y') if s.date else '',
            s.clientCode or '',
            s.clientName or '',
            service_label,
            s.get_requestType_display() or '',
            s.plan.label if s.plan else '',
            f'{float(s.total):.2f}',
            s.createdBy.name if s.createdBy else '',
        ]
        x = padding
        for i, v in enumerate(values):
            draw.text((x + 6, y + 10), str(v), fill='#1E293B', font=font)
            x += col_widths[i]
        y += row_height

    # Footer
    draw.text((padding, y + 10), f'{len(sales)} registros', fill='#94A3B8', font=font)

    buf = BytesIO()
    img.save(buf, format='PNG', optimize=True)
    buf.seek(0)
    return buf
